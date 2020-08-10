"""
Function: copy the formats of the excel cells into the current test case's excel file
Author: Xinran Wang
Date: 08/05/2020
"""

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy
import time

def acc_convert_function(column_letter, row_num, data_type, coeff_dic):
    """
    Fill in the ACC calculation formulas for assigned data type name
    :param column_letter: the column letter in which the original numeric data will be located
    :param row_num: the row number of the current cell we are processing
    :param data_type: a string representing the data type name
    :param coeff_dic: a dictionary storing the data type name & formula coefficient relationship
    :return: the calculation formula to be filled into the cell (None if such cell doesn't exist)
    """
    data_no = data_type[:4]
    coefficient = coeff_dic[data_no]
    try:
        return "=IF(HEX2DEC(" + column_letter + str(row_num) + ")>=32768,-HEX2DEC(DEC2HEX(65536-HEX2DEC(" + column_letter +\
           str(row_num) + ")))*" + str(coefficient) + ",HEX2DEC(" + column_letter + str(row_num) + ")*" +\
           str(coefficient) + ")"
    except:
        return None

def process_acc_data(temp_dir, file_dir, data_file_name, acc_column, coeff_dic, name_to_description, pbar, progressbar_hist, progressbar_length):
    """
    Copy the formats of template based on the number of acc data, make changes directly on the acc data sheet of the excel file
    :param temp_dir: the directory of template excel file
    :param file_dir: the directory where the current data excel locates
    :param data_file_name: the test data excel file we want to process
    :param acc_column: a list containing columns of ACC data in the ACC dataframe
    :param coeff_dic: a dictionary storing the data type name & formula coefficient relationship
    :param name_to_description: a dictionary storing the data type name & description texts relationship
    :param pbar: a progressbar object
    :param progressbar_hist: the length of progressbar before this execution
    :param progressbar_length: the length of progressbar for processing this part of data
    :return: None
    """
    wb = load_workbook(temp_dir)
    acc_template = wb[wb.sheetnames[0]]
    acc_rows = acc_template.max_row
    acc_columns = acc_template.max_column

    book_path = os.path.join(file_dir, data_file_name)
    book = load_workbook(book_path)
    acc_sheet = book.create_sheet("Acc_Data", 0)
    # acc_sheet = book["Acc_Data"]

    acc_prev_sheet_cols = 0

    total_len = len(acc_column)
    unit_len = progressbar_length / total_len

    for index, obj in enumerate(acc_column):

        print("Creating slots for:", obj)
        update_pbar(pbar, progressbar_hist + unit_len * (index + 1))

        if index == 0:
            low_bound = 0
        else:
            low_bound = 1
        for rows in range(acc_rows):
            acc_sheet.row_dimensions[rows + 1].height = acc_template.row_dimensions[rows + 1].height
            for col in range(low_bound, acc_columns):
                acc_sheet.column_dimensions[get_column_letter(acc_prev_sheet_cols + col + 1 - low_bound)].width = \
                    acc_template.column_dimensions[get_column_letter(col + 1)].width
                cell_val = acc_template.cell(rows + 1, col + 1).value
                if cell_val == "Content":
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound, value=obj+name_to_description[obj[:4]])
                elif type(cell_val) == str and cell_val.startswith("="):
                    col_idx = get_column_letter(acc_prev_sheet_cols + col - low_bound)
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound,
                                   value=acc_convert_function(col_idx, rows + 1, obj, coeff_dic))
                else:
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound, value=cell_val)
                if acc_template.cell(rows + 1, col + 1).has_style:
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound).font = copy(
                        acc_template.cell(rows + 1, col + 1).font)
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound).border = copy(
                        acc_template.cell(rows + 1, col + 1).border)
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound).fill = copy(
                        acc_template.cell(rows + 1, col + 1).fill)
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound).number_format = copy(
                        acc_template.cell(rows + 1, col + 1).number_format)
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound).protection = copy(
                        acc_template.cell(rows + 1, col + 1).protection)
                    acc_sheet.cell(row=rows + 1, column=acc_prev_sheet_cols + col + 1 - low_bound).alignment = copy(
                        acc_template.cell(rows + 1, col + 1).alignment)
        acc_prev_sheet_cols += (3 - low_bound)

        to_merge_end = get_column_letter(acc_prev_sheet_cols)
        to_merge_start = get_column_letter(acc_prev_sheet_cols - 1)
        acc_sheet.merge_cells(to_merge_start + "2:" + to_merge_end + "2")

    acc_sheet.merge_cells("B1:" + to_merge_end + "1")

    try:
        book.remove(book["Sheet"])
    except:
        pass

    wb.close()
    book.save(book_path)
    book.close()


def edr_convert_function(col_letter, text):
    """
    Fill in the EDR calculation formulas
    :param col_letter: the column letter in which the original numeric data will be located
    :param text: the formula text for this row of data
    :return: the calculation formula to be filled into the cell
    """
    parenthesis_idx = text.find("(")
    return text[:parenthesis_idx+1] + col_letter + text[parenthesis_idx+2:]

def process_edr_data(temp_dir, file_dir, data_file_name, edr_column, pbar, progressbar_hist, progressbar_length):
    """
    Copy the formats of template based on the number of edr data, make changes directly on the edr data sheet of the excel file
    :param temp_dir: the directory of template excel file
    :param file_dir: the directory where the current data excel locates
    :param data_file_name: the test data excel file we want to process
    :param edr_column: a list containing columns of ACC data in the EDR dataframe
    :param pbar: a progressbar object
    :param progressbar_hist: the length of progressbar before this execution
    :param progressbar_length: the length of progressbar for processing this part of data
    :return: None
    """
    wb = load_workbook(temp_dir)

    edr_template = wb[wb.sheetnames[1]]
    edr_rows = edr_template.max_row
    edr_columns = edr_template.max_column

    book_path = os.path.join(file_dir, data_file_name)
    book = load_workbook(book_path)
    edr_sheet = book.create_sheet("EDR", 1)
    # edr_sheet = book["EDR"]

    edr_prev_sheet_cols = 0

    # process merged cells
    wm = list(edr_template.merged_cells)
    edr_merged = []
    if len(wm) > 0:
        for i in range(0, len(wm)):
            cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
            edr_merged.append(cell2)

    total_len = len(edr_column)
    unit_len = progressbar_length / total_len

    for index, obj in enumerate(edr_column):

        print("Creating slots for:", obj)
        update_pbar(pbar, progressbar_hist + unit_len * (index+1))

        for rows in range(edr_rows):
            edr_sheet.row_dimensions[rows + 1].height = edr_template.row_dimensions[rows + 1].height
            for col in range(edr_columns):
                edr_sheet.column_dimensions[get_column_letter(edr_prev_sheet_cols + col + 1)].width = \
                    edr_template.column_dimensions[get_column_letter(col + 1)].width
                cell_val = edr_template.cell(rows + 1, col + 1).value
                if cell_val == "Test Case Name":
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1, value=obj)
                elif type(cell_val) == str and cell_val.startswith("="):
                    data_idx = get_column_letter(edr_prev_sheet_cols + col)
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1, value=edr_convert_function(data_idx, cell_val))
                else:
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1, value=cell_val)
                if edr_template.cell(rows + 1, col + 1).has_style:
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1).font = copy(
                        edr_template.cell(rows + 1, col + 1).font)
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1).border = copy(
                        edr_template.cell(rows + 1, col + 1).border)
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1).fill = copy(
                        edr_template.cell(rows + 1, col + 1).fill)
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1).number_format = copy(
                        edr_template.cell(rows + 1, col + 1).number_format)
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1).protection = copy(
                        edr_template.cell(rows + 1, col + 1).protection)
                    edr_sheet.cell(row=rows + 1, column=edr_prev_sheet_cols + col + 1).alignment = copy(
                        edr_template.cell(rows + 1, col + 1).alignment)
        edr_prev_sheet_cols += 5

        a_col = get_column_letter(edr_prev_sheet_cols - 4)
        b_col = get_column_letter(edr_prev_sheet_cols - 3)
        c_col = get_column_letter(edr_prev_sheet_cols - 2)
        d_col = get_column_letter(edr_prev_sheet_cols - 1)
        for cell in edr_merged:
            new = cell.replace("A", a_col).replace("B", b_col).replace("C", c_col).replace("D", d_col)
            edr_sheet.merge_cells(new)

    try:
        book.remove(book["Sheet"])
    except:
        pass

    wb.close()
    book.save(book_path)
    book.close()

def update_pbar(progress_bar, progress):
    """
    Updating the progressbar object
    :param progress_bar: the progressbar object
    :param progress: a number representing the current progress
    :return: None
    """
    progress_bar.update(progress)
    time.sleep(0.01)
