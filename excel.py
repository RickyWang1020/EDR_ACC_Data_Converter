"""
Function: operations of creating excel file and filling data into excel
Author: Xinran Wang
Date: 08/11/2020
"""

import openpyxl
import os
from openpyxl import Workbook


def create_folder(directory, name):
    """
    Create a folder in the assigned directory with assigned name
    :param directory: the directory where the new folder will be located
    :param name: the folder's name
    :return: the absolute path of the folder (None if the folder already exists under the directory)
    """
    path = os.path.join(directory, name)
    exist = os.path.exists(path)
    if exist:
        print("The folder: " + path + " already exists, cannot create new folder!!")
        return None
    else:
        os.mkdir(path)
        print("Created a new folder in: " + directory)
        return path

def create_excel(directory, file_name):
    """
    Creating an empty excel file in the assigned directory
    :param directory: the location of the excel that will be created
    :param file_name: the name of the excel that will be created
    :return: the full name of the excel file (with .xlsx)
    """
    wb = Workbook()
    new_file = file_name + ".xlsx"
    path = os.path.join(directory, new_file)
    wb.save(path)
    return new_file


def fill_excel_for_one_testcase(working_dir, FA_dataframe, B0_dataframe, file_name):
    """
    Fill the test data of one test case into the excel file
    :param working_dir: the directory where the excel file for current test case locates
    :param FA_dataframe: the pandas dataframe containing FA... data
    :param B0_dataframe: the pandas dataframe containing B0... data
    :param file_name: the file name of the current test case
    :return: None
    """
    absolute_path = os.path.join(working_dir, file_name)
    wb = openpyxl.load_workbook(absolute_path)

    # process B0 data
    try:
        b0 = wb["Acc_Data"]
        print("Filling in ACC Data for: " + file_name + "...\n")
        for row_position, col in enumerate(B0_dataframe.columns):
            column = B0_dataframe[col]
            for i in range(1, len(column)+1):
                obj = column[i-1]
                b0.cell(row=3+i, column=(1+row_position)*2).value = obj
    except:
        print("No ACC Data to fill in " + file_name + "...\n")

    # process FA data
    try:
        fa = wb["EDR"]
        print("Filling in EDR Data for: " + file_name + "...\n")
        for r_p, c in enumerate(FA_dataframe.columns):
            colu = FA_dataframe[c]
            for j in range(1, len(colu)+1):
                item = colu[j-1][2:]
                fa.cell(row=3+j, column=3+5*r_p).value = item
    except:
        print("No EDR Data to fill in " + file_name + "...\n")
    wb.save(absolute_path)
